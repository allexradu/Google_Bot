from time import sleep
from urllib.error import HTTPError
from urllib.request import urlretrieve

import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver


def add_data_to_excel(d_images, cell_letters, is_photo):
    try:
        for cell_letter in cell_letters:
            if is_photo:
                key = 'Image {column_image_number}'.format(column_image_number = cell_letters.index(cell_letter) + 1)
            else:
                key = 'Image {column_image_number} Link'.format(
                    column_image_number = cell_letters.index(cell_letter) + 1)

            column_rows = d_images[key]

            df_new = pd.DataFrame.from_records({key: column_rows})

            wb = load_workbook('table.xlsx')

            ws = wb['Sheet1']

            for cell_row_number, row in df_new.iterrows():
                cell = '%(column_letter)s%(cell_row_number)s' % dict(column_letter = cell_letter,
                                                                     cell_row_number = cell_row_number + 2)
                ws[cell] = row[0]

            wb.save('table.xlsx')
    except:
        print('WARNING!! Data NOT written to Excel. Make sure you keep your workbook CLOSED!')


class GoogleBot:
    def __init__(self, query, file_base, delay, number_of_images, down_images, main_index, link_i):
        # Activating the Chrome Driver
        self.link_index = link_i
        self.filename = ''
        self.driver = webdriver.Chrome()
        url = 'https://www.google.com/search?q={q}'.format(q = query)

        # Navigating to the Google Search for particular product
        self.driver.get(url)

        # Navigating to the Google Images tab
        self.driver.find_element_by_xpath(
            '//*[@id="hdtb-msb-vis"]/div[2]/a') \
            .click()

        sleep(delay)

        self.thumbnails_xpath = []

        # Adding the x paths of the thumbnails to the list
        for ind in range(number_of_images):
            self.thumbnails_xpath.append(
                '//*[@id="islrg"]/div[1]/div[{index}]/a[1]/div[1]/img'.format(index = ind + 1))

        self.larger_image_xpath = '//*[@id="Sva75c"]/div/div/div[3]/div[2]/div/div[1]/div[1]/div/div[2]/a/img'

        # Downloading the images
        for idx in range(number_of_images):
            self.image_download(delay = delay, file_base = file_base, query = query,
                                thumbnail = self.thumbnails_xpath[idx],
                                image = self.larger_image_xpath,
                                image_index = idx + 1,
                                d_load_images = down_images,
                                m_index = main_index)
        photo_cells = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        add_data_to_excel(d_images = down_images, cell_letters = photo_cells, is_photo = True)
        link_cells = ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
        add_data_to_excel(d_images = down_images, cell_letters = link_cells, is_photo = False)

        self.driver.close()

    def image_download(self, delay, file_base, query, thumbnail, image, image_index, d_load_images, m_index,
                       lnk_index):

        def excel_add_data_when_no_error():
            key = 'Image %(column_image_number)s' % dict(column_image_number = image_index)
            d_load_images[key].insert(m_index, self.filename)

        def excel_add_na_when_error():
            key_err = 'Image %(column_image_number)s' % dict(column_image_number = image_index)
            d_load_images[key_err].insert(m_index, 'n/a')

        try:
            # Navigating to the 1st thumbnail of the search and CLICKing in it
            self.driver.find_element_by_xpath(thumbnail) \
                .click()
            sleep(delay)

            # Finding the larger image and assigning its tag to the img variable
            img = self.driver.find_element_by_xpath(image)

            # Getting the URL out the img attribute
            src = img.get_attribute('src')

            self.filename = file_base + '{image_index}.jpg'.format(image_index = image_index)

            # Downloading the (bigger) image and storing it locally
            urlretrieve(src, self.filename)
        except FileNotFoundError as err:
            excel_add_na_when_error()
            print(err)  # something wrong with local path
        except HTTPError as err:
            excel_add_na_when_error()
        except:
            excel_add_na_when_error()

            # something unexpected went wrong
            print(
                'Unknown Error Image {image_index} - {q}'.format(q = query, image_index = image_index))

        else:
            print(lnk_index)
            # Adding filename to download_images to be saved in Excel
            excel_add_data_when_no_error()
            # Letting the user know that the 1st image of the first product has been downloaded
            print('{q} - {filename} downloaded'.format(q = query, filename = self.filename))


print('========================================================================')
print('|           Google_Bot 1.0.1.4 by Allex Radu [www.ATFR.net]             |')
print('|     Get the latest version at https://github.com/allexradu/gBot       |')
print('========================================================================')
print('| Instructions: Save your Excel Workbook as "a.xls" and place it in     |')
print('| the same folder as this file, make sure the file in not opened.       |')
print('========================================================================')
print('|      WARNING!!! WRITE THIS DOWN! To stop the bot press CTRL + C       |')
print('========================================================================')

no_of_images = 0

link_index = 0

downloaded_images = {'Image 1': [''], 'Image 2': [''], 'Image 3': [''], 'Image 4': [''],
                     'Image 5': [''], 'Image 6': [''], 'Image 7': [''], 'Image 8': [''], 'Image 9': [''],
                     'Image 1 Link': [], 'Image 2 Link': [], 'Image 3 Link': [], 'Image 4 Link': [], 'Image 5 Link': [],
                     'Image 6 Link': [], 'Image 7 Link': [], 'Image 8 Link': [], 'Image 9 Link': []}

while True:
    try:
        no_of_images = int(input('Number of images per product [1-9]: '))
    except:
        print('Invalid Input!!! Try again!')
    else:
        if not (0 < no_of_images <= 9):
            print('Number not in range, try again!!')
            continue
        else:
            break

while True:
    try:
        seconds_delay = float(input('Number of seconds delay from one image to another: ' +
                                    '\n (the slower the computer / connection the higher the number)' +
                                    '\n [Minimum 1 sec recommended] seconds: '))
    except:
        print('Invalid Input!!! Try again!')
    else:
        break

# Reading first column of a local excel file
try:
    df = pd.read_excel('a.xls', sheet_name = 0)
    print('Excel Read Complete!')

    product_names = df['Name'].tolist()

    for i in range(len(product_names)):
        print(product_names[i])

    # Placing all the product names in a list

    for i in range(len(product_names)):
        if i == 0:
            link_index = 2
        else:
            link_index = i + 1
        file_name_base = 'A' + '{num}'.format(num = (100 + i))
        product_name = product_names[i]
        GoogleBot(product_name, file_name_base, seconds_delay, no_of_images, downloaded_images, i, link_index)
except:
    print('Excel File NOT READ. Name your file "a.xls" with the first column "Name"')
    print('and place it in the same directory and the bot file.')
