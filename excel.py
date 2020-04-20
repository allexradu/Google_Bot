import pandas as pd
import platform
from openpyxl import load_workbook

downloaded_images = {'Image 1': [''], 'Image 2': [''], 'Image 3': [''], 'Image 4': [''],
                     'Image 5': [''], 'Image 6': [''], 'Image 7': [''], 'Image 8': [''], 'Image 9': ['']}

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else '../excel/a.xlsx'


def read_excel_first_column():
    # Reading first column of a local excel file
    try:
        if platform.system() == 'Windows':
            df = pd.read_excel(table_location, sheet_name = 0)
        else:
            df = pd.read_excel(table_location, sheet_name = 0)

        print('Excel Read Complete!')

        product_names = df['Name'].tolist()

        for i in range(len(product_names)):
            print(product_names[i])
        return product_names
    except:
        print('Excel File NOT READ. Name your file "a.xls" with the first column "Name"')
        print('and place it in the same directory and the bot file.')


def add_data_to_excel():
    cell_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    try:
        for i in range(9):
            key = 'Image {column_image_number}'.format(column_image_number = i + 1)

            column_rows = downloaded_images[key]

            df_new = pd.DataFrame.from_records({key: column_rows})

            wb = load_workbook(table_location)

            ws = wb['Sheet1']

            for cell_row_number, row in df_new.iterrows():
                cell = '%(column_letter)s%(cell_row_number)s' % dict(column_letter = cell_letters[i],
                                                                     cell_row_number = cell_row_number + 2)
                ws[cell] = row[0]

            wb.save(table_location)
    except:
        print('WARNING!! Data NOT written to Excel. Make sure you keep your workbook CLOSED!')
